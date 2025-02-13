import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font
import warnings
import torch
import torch.nn as nn
import torch.optim as optim
from sklearn.preprocessing import MinMaxScaler

plt.rcParams['figure.dpi'] = 100
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False
warnings.filterwarnings("ignore")

def run_lstm(file_path, data_column, pre_num,df):
    # 集中管理可调节参数
    config = {
        'window_size': 6,
        'num_epochs': 200,
        'hidden_size': 50,
        'learning_rate': 0.001,
        'weight_decay': 1e-6  # 添加正则化系数
    }

    time_series = df[data_column].values.reshape(-1, 1)
    # 数据归一化
    scaler = MinMaxScaler()
    scaled_series = scaler.fit_transform(time_series)

    # 划分训练集和测试集
    train_size = int(len(scaled_series) * 0.7)
    train_data = scaled_series[:train_size]
    test_data = scaled_series[train_size:]

    # 创建训练和测试数据集
    X_train, y_train = create_dataset(train_data, config['window_size'])
    X_test, y_test = create_dataset(test_data, config['window_size'])

    # 转换数据形状
    X_train = X_train.reshape(-1, config['window_size'], 1)
    y_train = y_train.reshape(-1, 1)
    X_test = X_test.reshape(-1, config['window_size'], 1)
    y_test = y_test.reshape(-1, 1)

    # 转换为PyTorch张量
    X_train_tensor = torch.tensor(X_train, dtype=torch.float32)
    y_train_tensor = torch.tensor(y_train, dtype=torch.float32)
    X_test_tensor = torch.tensor(X_test, dtype=torch.float32)
    y_test_tensor = torch.tensor(y_test, dtype=torch.float32)

    # 定义模型
    model = LSTMModel(input_size=1, hidden_size=config['hidden_size'], output_size=1, num_layers=2)

    # 定义损失函数和优化器
    criterion = nn.MSELoss()
    optimizer = optim.Adam(model.parameters(),
                           lr=config['learning_rate'],
                           weight_decay=config['weight_decay'])  # 添加正则化

    # 训练模型
    for epoch in range(config['num_epochs']):
        model.train()
        optimizer.zero_grad()
        outputs, _ = model(X_train_tensor)
        loss = criterion(outputs, y_train_tensor)
        loss.backward()
        optimizer.step()
        if (epoch + 1) % 10 == 0:
            print(f"Epoch [{epoch + 1}/{config['num_epochs']}], Loss: {loss.item():.4f}")

    # 测试模型
    model.eval()
    with torch.no_grad():
        train_predict, _ = model(X_train_tensor)
        test_predict, _ = model(X_test_tensor)

    # 反归一化
    train_predict = scaler.inverse_transform(train_predict.numpy())
    test_predict = scaler.inverse_transform(test_predict.numpy())

    # 生成预测结果
    original_data = df[data_column].values
    total_samples = len(train_predict) + len(test_predict)

    # 将拟合值和预测值填充到原数据中
    fitted_values_padded = np.full_like(original_data, np.nan, dtype=np.float64)
    fitted_values_padded[config['window_size']:config['window_size'] + len(train_predict)] = train_predict.flatten()
    fitted_values_padded[
    train_size + config['window_size']:train_size + config['window_size'] + len(test_predict)] = test_predict.flatten()

    # 计算MSE
    valid_indices = ~np.isnan(fitted_values_padded)
    aligned_original = original_data[valid_indices]
    aligned_fitted = fitted_values_padded[valid_indices]
    mse = ((aligned_original - aligned_fitted) ** 2).mean()
    print(f"\n原始数据与拟合数据的MSE：{mse:.4f}")

    # 生成未来预测
    future_predictions = predict_future(model, pre_num, X_test_tensor[-1], config['window_size'], scaler)

    # 保存结果到Excel
    output_path = Path(file_path).parent / "LSTM_res.xlsx"
    combined_data = pd.DataFrame({
        'Original Data': original_data,
        'Fitted Values': fitted_values_padded
    })

    forecast_index = range(len(original_data), len(original_data) + pre_num)
    forecast_df = pd.DataFrame({'Forecast Values': future_predictions.flatten()}, index=forecast_index)
    combined_data = pd.concat([combined_data, forecast_df], axis=0)

    combined_data.to_excel(output_path, index=True, index_label="Index")

    # 加粗样式
    wb = load_workbook(output_path)
    ws = wb.active
    forecast_col = combined_data.columns.get_loc('Forecast Values') + 1  # 列数从1开始
    start_row = len(original_data) + 2  # 行从1开始
    for row in range(start_row, start_row + pre_num):
        ws.cell(row=row, column=forecast_col).font = Font(bold=True)
    wb.save(output_path)

    print(f"预测结果已保存至：{output_path}")
    return combined_data, mse, future_predictions
# 创建数据集
def create_dataset(data, window_size=3):
    X, y = [], []
    for i in range(len(data) - window_size):
        X.append(data[i:i + window_size])
        y.append(data[i + window_size])
    return np.array(X), np.array(y)

# 生成未来预测值
def predict_future(model, steps, last_input, window_size, scaler):
    model.eval()
    future = []
    current_input = last_input.numpy().flatten()
    with torch.no_grad():
        for _ in range(steps):
            input_tensor = torch.tensor(current_input[-window_size:], dtype=torch.float32).view(1, window_size, 1)
            pred, _ = model(input_tensor)
            future.append(pred.item())
            current_input = np.append(current_input, pred.item())
    return scaler.inverse_transform(np.array(future).reshape(-1, 1))

# 定义 LSTM 模型
class LSTMModel(nn.Module):
    def __init__(self, input_size, hidden_size, output_size, num_layers=2):
        super().__init__()
        self.hidden_size = hidden_size
        self.num_layers = num_layers
        self.lstm = nn.LSTM(input_size, hidden_size, num_layers, batch_first=True)
        self.fc = nn.Linear(hidden_size, output_size)

    def forward(self, x):
        h0 = torch.zeros(self.num_layers, x.size(0), self.hidden_size).requires_grad_()
        c0 = torch.zeros(self.num_layers, x.size(0), self.hidden_size).requires_grad_()
        out, (hn, cn) = self.lstm(x, (h0.detach(), c0.detach()))
        out = self.fc(out[:, -1, :])  # 只取最后一个时间步的输出
        return out, (hn, cn)
