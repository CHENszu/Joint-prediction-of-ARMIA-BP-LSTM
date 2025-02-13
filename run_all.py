import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font
import warnings

warnings.filterwarnings("ignore")

# 导入三个模型的函数（需要将原代码封装为函数）
from pre import run_arima
from pre1 import run_bp_nn
from pre2 import run_lstm

plt.rcParams['figure.dpi'] = 100
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False


def main():
    # 读取数据
    # 用户输入数据路径
    file_path = input("请输入数据文件路径（支持CSV和Excel文件）：")

    # 根据文件后缀自动判断文件类型
    file_ext = Path(file_path).suffix.lower()
    if file_ext == '.csv':
        df = pd.read_csv(file_path)
    elif file_ext in ('.xls', '.xlsx'):
        df = pd.read_excel(file_path)
    else:
        raise ValueError("不支持的文件类型，请使用CSV或Excel文件")

    data_column = input("请输入数据列的列名：")
    pre_num = int(input("请输入往后预测的步数："))
    if pre_num <= 0:
        print("预测步数必须大于0。")
        return

    # 运行三个模型
    print("\n=== 运行ARIMA模型 ===")
    df_arima, mse1, pre1 = run_arima(file_path, data_column, pre_num, df)

    print("\n=== 运行BP神经网络 ===")
    df_bp, mse2, pre2 = run_bp_nn(file_path, data_column, pre_num, df)

    print("\n=== 运行LSTM模型 ===")
    df_lstm, mse3, pre3 = run_lstm(file_path, data_column, pre_num, df)

    # 计算权重
    weights = np.array([1 / mse1, 1 / mse2, 1 / mse3])
    total_weight = weights.sum()
    normalized_weights = weights / total_weight

    # 将pre1转换为numpy数组并展平
    arima_values = pre1['Forecast Values'].to_numpy().flatten()

    # 将pre2和pre3展平为1D数组
    bp_values = pre2.flatten()
    lstm_values = pre3.flatten()

    # 验证形状一致性
    assert arima_values.shape == (pre_num,), f"ARIMA形状错误: {arima_values.shape}"
    assert bp_values.shape == (pre_num,), f"BPNN形状错误: {bp_values.shape}"
    assert lstm_values.shape == (pre_num,), f"LSTM形状错误: {lstm_values.shape}"

    # 重新计算最终预测值
    final_pre_values = (
            arima_values * normalized_weights[0] +
            bp_values * normalized_weights[1] +
            lstm_values * normalized_weights[2]
    )

    # 合并结果
    combined_df = pd.DataFrame({
        'ARIMA预测': arima_values,
        'BPNN预测': bp_values,
        'LSTM预测': lstm_values,
        '组合预测': final_pre_values
    })

    # === 新增代码：保存结果 ===
    output_path = Path(file_path).parent / "forest_result.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        combined_df.to_excel(writer, index=False, sheet_name="组合预测结果")

    # 设置标题行格式
    wb = load_workbook(output_path)
    ws = wb.active
    bold_font = Font(bold=True)

    # 设置第一行的字体为加粗
    for cell in ws[1]:
        cell.font = bold_font

    wb.save(output_path)
    print(f"\n预测结果已保存至：{output_path}")

    # === 新增代码：绘制对比图 ===
    plt.figure(figsize=(10, 6))
    original_values = df[data_column]

    # 绘制原始数据
    plt.plot(original_values.values, label="原始数据", color="blue", marker="o")

    # 生成预测数据的索引位置
    forecast_start = len(original_values)
    forecast_end = forecast_start + pre_num - 1
    forecast_index = np.arange(forecast_start, forecast_end + 1)

    # 绘制预测数据
    plt.plot(forecast_index, final_pre_values, label="组合预测", color="red", linestyle="--", marker="s")

    # 添加图表元素
    plt.title("原始数据与组合预测对比")
    plt.xlabel("时间序列")
    plt.ylabel(data_column)
    plt.legend()
    plt.grid(True)

    # 突出显示预测区间
    plt.axvspan(forecast_start - 0.5, forecast_end + 0.5, facecolor="yellow", alpha=0.3)

    # 自动调整刻度显示
    plt.xticks(np.arange(0, forecast_end + 2, max(1, (forecast_end + 2) // 10)))

    # 显示并保存图表
    # plot_path = Path(file_path).parent / "forecast_comparison.png"
    # plt.savefig(plot_path, dpi=300, bbox_inches="tight")
    # print(f"对比图表已保存至：{plot_path}")
    plt.show()

    # 输出权重信息
    print("\n=== 模型权重 ===")
    print(f"ARIMA权重：{normalized_weights[0]:.2%} (MSE={mse1:.4f})")
    print(f"BPNN权重：{normalized_weights[1]:.2%} (MSE={mse2:.4f})")
    print(f"LSTM权重：{normalized_weights[2]:.2%} (MSE={mse3:.4f})")

if __name__ == "__main__":
    main()