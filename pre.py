import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from statsmodels.stats.diagnostic import acorr_ljungbox
from pmdarima import auto_arima
from statsmodels.tsa.arima.model import ARIMA
import warnings
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font  # 导入Font类用于加粗

warnings.filterwarnings("ignore")

# 设置图片清晰度和中文字体
plt.rcParams['figure.dpi'] = 100
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False


def run_arima(file_path,data_column,pre_num,df):

    time_series = df.loc[:, data_column]
    # 数据检验
    print("\n=== 数据检验 ===")
    lb_p_value = test_white_noise(time_series)
    print(f"\n白噪声检验的p值：{lb_p_value:.4f}")
    if lb_p_value > 0.05:
        print("数据是白噪声，无法使用ARIMA模型进行预测。")
        return

    # 自动选择ARIMA参数
    print("\n=== 模型参数选择 ===")
    model_order = find_arima_order(time_series)
    p, d, q = model_order
    print(f"\n最优的ARIMA模型参数为：(p={p}, d={d}, q={q})")

    # 模型训练与预测
    model = ARIMA(time_series, order=(p, d, q))
    model_fit = model.fit()

    # 计算并输出MSE
    fitted_values = model_fit.fittedvalues
    original_aligned = time_series.loc[fitted_values.index]
    mse = ((original_aligned - fitted_values) ** 2).mean()
    print(f"\n原始数据与拟合数据的MSE：{mse:.4f}")

    # 输出模型详情
    print("\n=== 模型参数详情 ===")
    print(model_fit.summary())

    # 生成预测结果
    forecast = model_fit.forecast(steps=pre_num)
    forecast_index = np.arange(len(time_series), len(time_series) + pre_num)
    forecast_series = pd.Series(forecast, index=forecast_index)

    # 保存结果到Excel
    output_path = Path(file_path).parent / "ARIMA_res.xlsx"
    # 保存原始数据和拟合数据为DataFrame
    fitted_values = model_fit.fittedvalues.rename('Fitted Values')
    original_data = df[data_column].rename('Original Data')
    combined_df = pd.concat([original_data, fitted_values], axis=1)

    # 将预测值追加到DataFrame的末尾
    forecast_df = forecast_series.to_frame(name='Forecast Values')
    combined_df = pd.concat([combined_df, forecast_df])

    # 保存为Excel文件
    combined_df.to_excel(output_path, index=True, index_label="Index")

    # 使用openpyxl加载Excel文件并设置样式
    wb = load_workbook(output_path)
    ws = wb.active

    # 遍历预测数据部分（最后一部分pre_num行）
    start_row = len(time_series) + 2  # 从原始数据的最后一行开始
    end_row = start_row + pre_num - 1
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=ws.max_column)  # 最后一列是预测值
        cell.font = Font(bold=True)  # 设置字体加粗

    # 保存修改后的Excel文件
    wb.save(output_path)
    print(f"预测结果已保存至：{output_path}")

    return combined_df, mse, forecast_df

def test_white_noise(series, lags=10):
    """白噪声检验（Ljung-Box检验）"""
    Q = acorr_ljungbox(series, lags=[lags], return_df=True)
    return Q['lb_pvalue'].iloc[0]


def find_arima_order(series):
    """自动选择ARIMA参数"""
    model = auto_arima(
        series,
        seasonal=False,
        d=None,
        trace=True,
        error_action='ignore',
        suppress_warnings=True
    )
    return model.order
