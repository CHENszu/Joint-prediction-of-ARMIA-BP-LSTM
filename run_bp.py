import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import warnings
import torch
import torch.nn as nn
import torch.optim as optim
from sklearn.preprocessing import MinMaxScaler

plt.rcParams['figure.dpi'] = 100
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False
warnings.filterwarnings("ignore")


def run_bp_nn():
    # 集中管理可调节参数
    config = {
        'window_size': 3,
        'num_epochs': 200,
        'hidden_size_1': 80,
        'hidden_size_2': 50,
        'learning_rate': 0.001,
        'weight_decay': 1e-6  # 添加正则化系数
    }

    # 输入数据文件路径
    file_path = input("请输入数据文件路径（支持CSV和Excel文件）：")
    file_ext = Path(file_path).suffix.lower()

    # 根据文件类型读取数据
    if file_ext == '.csv':
        df = pd.read_csv(file_path)
    elif file_ext in ('.xls', '.xlsx'):
        df = pd.read_excel(file_path)
    else:
        raise ValueError("不支持的文件类型，请使用CSV或Excel文件")

    # 输入数据列名
    data_column = input("请输入数据列的列名：")
    time_series = df[data_column].values.reshape(-1, 1)

    print("\n=== 预测与绘图 ===")
    pre_num = int(input("请输入要预测的步数："))
    if pre_num <= 0:
        print("预测步数必须大于0。")
        return

    # 数据归一化
    scaler = MinMaxScaler()
    scaled_series = scaler.fit_transform(time_series)

    # 划分训练集和测试集
    train_size = int(len(scaled_series) * 0.7)
    train_data = scaled_series[:train_size]
    test_data = scaled_series[train_size:]

    # 创建训练和测试数据集
    X_train, y_train = create_dataset(train_data, config['window_size'])
    X_test, y_test = create_dataset(test_data, config['window_size'])

    # 转换数据形状
    X_train = X_train.reshape(-1, config['window_size'])
    y_train = y_train.reshape(-1, 1)
    X_test = X_test.reshape(-1, config['window_size'])
    y_test = y_test.reshape(-1, 1)

    # 转换为PyTorch张量
    X_train_tensor = torch.tensor(X_train, dtype=torch.float32)
    y_train_tensor = torch.tensor(y_train, dtype=torch.float32)
    X_test_tensor = torch.tensor(X_test, dtype=torch.float32)
    y_test_tensor = torch.tensor(y_test, dtype=torch.float32)

    # 定义模型
    model = BPNeuralNetwork(input_size=config['window_size'],
                            hidden_size_1=config['hidden_size_1'],
                            hidden_size_2=config['hidden_size_2'],
                            output_size=1)

    # 定义损失函数和优化器
    criterion = nn.MSELoss()
    optimizer = optim.Adam(model.parameters(),
                           lr=config['learning_rate'],
                           weight_decay=config['weight_decay'])  # 添加正则化

    # 训练模型
    for epoch in range(config['num_epochs']):
        model.train()
        optimizer.zero_grad()
        outputs = model(X_train_tensor)
        loss = criterion(outputs, y_train_tensor)
        loss.backward()
        optimizer.step()
        if (epoch + 1) % 10 == 0:
            print(f"Epoch [{epoch + 1}/{config['num_epochs']}], Loss: {loss.item():.4f}")

    # 测试模型
    model.eval()
    with torch.no_grad():
        train_predict = model(X_train_tensor).numpy()
        test_predict = model(X_test_tensor).numpy()

    # 反归一化
    train_predict = scaler.inverse_transform(train_predict)
    test_predict = scaler.inverse_transform(test_predict)

    # 生成预测结果
    original_data = df[data_column].values
    total_samples = len(train_predict) + len(test_predict)

    # 将拟合值和预测值填充到原数据中
    fitted_values_padded = np.full_like(original_data, np.nan, dtype=np.float64)
    fitted_values_padded[config['window_size']:config['window_size'] + len(train_predict)] = train_predict.flatten()
    fitted_values_padded[
    train_size + config['window_size']:train_size + config['window_size'] + len(test_predict)] = test_predict.flatten()

    # 计算MSE
    valid_indices = ~np.isnan(fitted_values_padded)
    aligned_original = original_data[valid_indices]
    aligned_fitted = fitted_values_padded[valid_indices]
    mse = ((aligned_original - aligned_fitted) ** 2).mean()
    print(f"\n原始数据与拟合数据的MSE：{mse:.4f}")

    # 生成未来预测
    future_predictions = predict_future(model, pre_num, X_test_tensor[-1], config['window_size'], scaler)

    # 保存结果到Excel
    output_path = Path(file_path).parent / "BP_NN_res.xlsx"
    combined_data = pd.DataFrame({
        'Original Data': original_data,
        'Fitted Values': fitted_values_padded
    })

    forecast_index = range(len(original_data), len(original_data) + pre_num)
    forecast_df = pd.DataFrame({'Forecast Values': future_predictions.flatten()}, index=forecast_index)
    combined_data = pd.concat([combined_data, forecast_df], axis=0)

    combined_data.to_excel(output_path, index=True, index_label="Index")

    # 加粗样式
    wb = load_workbook(output_path)
    ws = wb.active
    forecast_col = combined_data.columns.get_loc('Forecast Values') + 1  # 列数从1开始
    start_row = len(original_data) + 2  # 行从1开始
    for row in range(start_row, start_row + pre_num):
        ws.cell(row=row, column=forecast_col).font = Font(bold=True)
    wb.save(output_path)

    print(f"预测结果已保存至：{output_path}")

    # 绘制结果
    plt.figure(figsize=(12, 6))
    plt.plot(original_data, label="原始数据", color="blue", alpha=0.6)
    plt.plot(np.where(~np.isnan(fitted_values_padded))[0], fitted_values_padded[~np.isnan(fitted_values_padded)],
             label="拟合数据", color="green", linewidth=1.5)
    plt.plot(range(len(original_data), len(original_data) + pre_num), future_predictions,
             label=f"{pre_num}步预测", color="red", linestyle="--", marker='o')
    plt.title("时间序列分析结果（BP神经网络）", fontsize=14)
    plt.xlabel("时间序列", fontsize=12)
    plt.ylabel("数值", fontsize=12)
    plt.grid(True, alpha=0.3)
    plt.legend()
    plt.tight_layout()
    plt.show()


# 创建数据集
def create_dataset(data, window_size=3):
    X, y = [], []
    for i in range(len(data) - window_size):
        X.append(data[i:i + window_size])
        y.append(data[i + window_size])
    return np.array(X), np.array(y)


# 生成未来预测值
def predict_future(model, steps, last_input, window_size, scaler):
    model.eval()
    future = []
    current_input = last_input.numpy().flatten()
    with torch.no_grad():
        for _ in range(steps):
            input_tensor = torch.tensor(current_input[-window_size:], dtype=torch.float32).view(1, -1)
            pred = model(input_tensor).numpy().flatten()
            future.append(pred[0])
            current_input = np.append(current_input, pred[0])
    return scaler.inverse_transform(np.array(future).reshape(-1, 1))


# 定义神经网络模型（增加一个隐藏层）
class BPNeuralNetwork(nn.Module):
    def __init__(self, input_size, hidden_size_1, hidden_size_2, output_size):
        super().__init__()
        self.fc1 = nn.Linear(input_size, hidden_size_1)
        self.relu = nn.ReLU()
        self.fc2 = nn.Linear(hidden_size_1, hidden_size_2)  # 新增隐藏层
        self.fc3 = nn.Linear(hidden_size_2, output_size)

    def forward(self, x):
        x = self.fc1(x)
        x = self.relu(x)
        x = self.fc2(x)
        x = self.relu(x)
        x = self.fc3(x)
        return x


if __name__ == "__main__":
    run_bp_nn()